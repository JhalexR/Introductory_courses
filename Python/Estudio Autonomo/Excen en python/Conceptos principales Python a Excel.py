# resumen de conceptos principales para trabajar en Python con Excel

# Librerias que se importan
import pandas as pd # para manejar bases de datos
import openpyxl # para manejar Excel


# Cargar un archivo Excel a una variable
# forma 1
file_path = 'descarga4.xlsx'
Tabs_monitor = pd.read_excel(file_path) # el archivo debe estar en el mismo lugar del archivo con el código

# forma eficiente dos 
# 1 metodo de carga base de datos desde archivo .CSV
fifa = pd.read_csv(r"C:\Users\JohnAlexanderPeñaloz\OneDrive - PayU\Principal\2024\Git\Introductory_courses\Python\Evoluciona desde Excel a Python\data.csv", sep=",")

# 2 metodo de carga base de datos desde archivo .xlsx
fifa_2 = pd.read_excel(r"C:\Users\JohnAlexanderPeñaloz\OneDrive - PayU\Principal\2024\Git\Introductory_courses\Python\Evoluciona desde Excel a Python\data.xlsx")

# Crear una tabla dinámica guardarla en una variable 
# ordenar una tabla dinámica

#tabla - numero de transacciones con el mismo BIN
trx_id_x_bin = Tabs_monitor.pivot_table(index='bin',values='transaccion_id',aggfunc='count')
trx_id_x_bin = trx_id_x_bin.sort_values(by='transaccion_id', ascending=False)
trx_id_x_bin = pd.DataFrame(trx_id_x_bin.reset_index())


# crear nuevo libro
# importar librerías 
from openpyxl import Workbook # para trabajar con libros de Excel
from openpyxl import load_workbook # para trabajar con libros de Excel
wb = Workbook() # variable del libro wb

# colocar nombre a hoja activa
ws = wb.active # variable de la hoja = ws
ws.title = "Datos vs TRX" #nombre de la hoja por transacciones

# formato de datos a columnas

# para formatos de Excel 
# importar librerías 
from openpyxl.styles import Font, PatternFill # para fuentes y formato de relleno
from openpyxl.styles import Alignment, Border, Side #para formato de celdas 

#formato de hoja 
ws.sheet_view.showGridLines = False # quitar líneas

# ajustar tamaño de columnas
ws.column_dimensions['B'].width = 2 * ws.column_dimensions['B'].width 
ws.column_dimensions['C'].width = 2 * ws.column_dimensions['C'].width


# colocar formato condicional escala de color
# importar librerías escala de color 
from openpyxl.formatting.rule import ColorScaleRule # formato condicional escalas de color 
# guardar formato en variable
color_scale_rule = ColorScaleRule(start_type='max', start_color='00FF00',
                                  mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                  end_type='min', end_color='FF0000')
# aplicar al rango de celdas
ws.conditional_formatting.add('C11:C30', color_scale_rule)

# escribir algo directamente en una celda 
ws['B4'] = "TOP 20"
ws['B6'] = "Cantidad de transacciones"


# colocar bordes a la celda
# primero se guarda el formato en una variable
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
# luego se aplica a la celdas en un ciclo

# formato de celdas fuente y relleno 
ws['C10'].font = Font(bold=True, size=12)
ws['B10'].fill = PatternFill(start_color="00E0FFFF", end_color="00E0FFFF", fill_type="solid")

# recorrer celdas con ciclos
# guardar la posición de las filas y columnas en variables 
col = 'B'
row = 11
cell = f"{col}{row}" # guardar la posición de la celda en variable

count_Fila = 0
# recorrer celdas con while usando las var anteriores y aplicando formato
while count_Fila < 20:
    ws[cell] = trx_id_x_bin.at[count_Fila, 'bin'] # colocando contenido de una variable tabla 
    ws[cell].border = thin_border # colcando el borde 
    ws[cell].number_format = '@' # colcando el formato de numero o texto 
    count_Fila += 1 # aumentar contador
    row += 1		# aumentar fila
    cell = f"{col}{row}" # guardar posición de la nueva celda

# graficos

#gráfico dispersión 
# importar librerías para grafico de dispersión 
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
chart_dis_trx_x_bin = ScatterChart()
chart_dis_trx_x_bin.style = 10
chart_dis_trx_x_bin.x_axis.title = 'BIN'
chart_dis_trx_x_bin.y_axis.title = 'Transacciones'

#Referencias datos -> grafico dispersion 
xvalues = Reference(ws, min_col=2, max_col=2, min_row=11, max_row=30)
yvalues = Reference(ws, min_col=3, max_col=3, min_row=10, max_row=30)

#Configuracion grafico dispersion 
series = Series(yvalues, xvalues, title_from_data=True)
chart_dis_trx_x_bin.series.append(series)
series.marker.symbol = "circle"
series.graphicalProperties.line.noFill = True
ws.add_chart(chart_dis_trx_x_bin, "E11")

#Grafico pie chart 
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

# configuración grafico de pie chart
chart_pie_trx_x_bin = PieChart()
labels = Reference(ws, min_col=2, min_row=11, max_row=30)
data = Reference(ws, min_col=3, min_row=10, max_row=30)
chart_pie_trx_x_bin.add_data(data, titles_from_data=True)
chart_pie_trx_x_bin.set_categories(labels)
ws.add_chart(chart_pie_trx_x_bin, "M11")

#Configuracion datalabel pie chart 
chart_pie_trx_x_bin.dataLabels = DataLabelList()
chart_pie_trx_x_bin.dataLabels.showPercent = True  # Mostrar porcentajes
chart_pie_trx_x_bin.dataLabels.showVal = False     # No mostrar valores
chart_pie_trx_x_bin.dataLabels.showCatName = False # No mostrar nombres de categorías
chart_pie_trx_x_bin.dataLabels.showLegendKey = False # No mostrar claves de leyenda

#Nueva Hoja 
#Creacion de hoja de Datos vs Datos 
ws2 = wb.create_sheet(title="Datos vs Datos", index=2)

#renombrando nueva hoja
ws2.sheet_view.showGridLines = False

# guardar libro de excel
output_path = 'Tabs_monitor.xlsx'
wb.save(output_path)
